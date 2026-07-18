import os
import re
import sys
import pickle
import shutil
import zipfile
import openpyxl
import warnings
import pythoncom
import configparser
import win32timezone
from pathlib import Path
from datetime import datetime
import win32com.client as win32
from collections import defaultdict
from openpyxl.styles import PatternFill, Font, Alignment

# ================= 配置与常量 =================

# ================= 获取安全运行目录 =================
def get_current_dir():
    # 获取程序当前所在的绝对目录，防打包后路径偏移
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))

BASE_DIR = get_current_dir()

# ================= 配置与常量 =================

DO_READ_FILE = False
DO_READ_FILE = True

CONFIG_PATH = os.path.join(BASE_DIR, 'config.ini')
config = configparser.ConfigParser()

# 如果没有配置文件，自动生成一个模板并提示

if not os.path.exists(CONFIG_PATH):
        config['SETTINGS'] = {
            'FILE_PATH': r'C:/Users/JYK/Downloads/CN Top Golden List.xlsx',
            'COPY_FILE_PATH_1': r'Z:/My documents/Games/Celeste/CN Top Golden List 备份',
            'COPY_FILE_PATH_2': r'Z:/My documents/Study/云盘/同步文件/CN Top Golden List 备份',
            'COPY_FILE_PATH_3': ''
        }
        with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
            config.write(f)
        print(f"未检测到配置文件，已在当前目录生成: {CONFIG_PATH}")
        print("请用记事本打开 config.ini，将配置修改为你电脑上对应的路径（注意路径中斜杠的方向），随后重新运行本程序。")
        print("FILE_PATH: 从腾讯文档下载的表格文件的路径。")
        print("COPY_FILE_PATH: 程序会将表格备份在本地的这个目录下，不需要备份的话留空即可。")
        print("程序每次读取表格数据之前，需要手动从腾讯文档下载表格文件，否则会报错。")
        input("按回车键退出...")
        sys.exit()

    # 读取配置
config.read(CONFIG_PATH, encoding='utf-8')
FILE_PATH = config.get('SETTINGS', 'FILE_PATH')
COPY_FILE_PATH_1 = config.get('SETTINGS', 'COPY_FILE_PATH_1')
COPY_FILE_PATH_2 = config.get('SETTINGS', 'COPY_FILE_PATH_2')
COPY_FILE_PATH_3 = ''
try:
    COPY_FILE_PATH_3 = config.get('SETTINGS', 'COPY_FILE_PATH_3')
except Exception as e:
    pass

COPY_FILE_PATH_LIST = [COPY_FILE_PATH_1,COPY_FILE_PATH_2,COPY_FILE_PATH_3]

PICKLE_FILENAME = "results.pkl"
HYPERLINK_FILENAME = "all_hyperlinks.txt"
MAIN_BANG_FILENAME = "主榜.xlsx"


DIFFICULTY_TYPES = [
    "Tier -1",
    "High T0", "Mid T0", "Low T0",
    "High T1", "Mid T1", "Low T1",
    "High T2", "Mid T2", "Low T2",
    "High T3", "Mid T3", "Low T3",
    "Tier 4", "Tier 5", "Tier 6", "Tier 7", "?"
]

TIER_GROUPS = [
    ("T-1", ["Tier -1"]),
    ("T0", ["High T0", "Mid T0", "Low T0"]),
    ("T1", ["High T1", "Mid T1", "Low T1"]),
    ("T2", ["High T2", "Mid T2", "Low T2"]),
    ("T3", ["High T3", "Mid T3", "Low T3"]),
    ("Farewell", []),
    ("T4", ["Tier 4"]),
    ("T5", ["Tier 5"]),
    ("T6", ["Tier 6"]),
    ("T7", ["Tier 7"]),
    ("Udt", ["?"])
]

HIGH_T0_CHALLENGES = []
IS_SUB_TIERED = ["T0", "T1", "T2", "T3"]

CHECK_CONFIGS = [
    {"sheet": "详细信息", "r": 3, "c": 7, "mode": "matrix"},
    {"sheet": "第九章金草莓", "r": 4, "c": 3, "mode": "col_only"},
    {"sheet": "其他挑战", "r": 3, "c": 7, "mode": "matrix"}
]

CHECK_COLUMNS_MAIN = [1, 3, 5, 7, 9, 11, 13, 15, 17, 19]
TIER_MAP_MAIN = {
    "Tier -1": (1,"D05DE9"),

    "High T0": (3, "F874C6"), 
    "Mid T0":  (3, "FF97D8"), 
    "Low T0":  (3, "FCB5E0"),

    "High T1": (5, "FF7B67"), 
    "Mid T1":  (5, "FF9989"), 
    "Low T1":  (5, "FCB6AB"),

    "High T2": (7, "FFC874"), 
    "Mid T2":  (7, "FFD595"), 
    "Low T2":  (7, "F8DCB2"),

    "High T3": (9, "FFEC87"), 
    "Mid T3":  (9, "FFEBB0"), 
    "Low T3":  (9, "FBF3CF"), 

    "Tier 4":  (11, "B0FF78"),
    "Tier 5":  (13, "85E191"),
    "Tier 6":  (15, "8FDEFF"),
    "Tier 7":  (17, "96A6FF"),
    "?":       (19, "D8D8D8")
}

TIER_MAP_MAIN_REV = {
    1: "T-1", 3: "T0", 5: "T1", 7: "T2", 9: "T3",
    11: "T4", 13: "T5", 15: "T6", 17: "T7", 19: "Udt"
    }




# ================= 辅助工具函数 =================

def int_to_excel_column(n):
    """将数字转换为 Excel 列号 (1->A, 28->AB)"""
    if not isinstance(n, int) or n <= 0:
        raise ValueError("输入必须是正整数")
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result


def clean_player_name(value):
    """清理玩家名：去掉结尾的 [FC]"""
    if isinstance(value, str) and value.endswith(' [FC]'):
        return value[:-5]
    return value

def extract_number_from_formula(formula):
    """提取公式中最后一个 $ 后的数字"""
    if not formula:
        return None
    # 查找最后一个 $
    last_dollar = formula.rfind('$')
    if last_dollar == -1:
        return None
    
    # 提取后续数字
    match = re.search(r'\d+', formula[last_dollar:])
    return int(match.group()) if match else None

def clean_map_name_this_is_a_function(text):
    """
    提取文本匹配的核心部分：
    1. 提取所有中括号内容，并过滤掉 [C], [FC], [C/FC]
    2. 输出所有小括号及其内容的列表
    3. 移除所有中括号及其内容
    4. 忽略所有空格
    返回: 纯净名称+所有中括号内容 , 指标(C,FC) , [所有小括号内容]
    """
    if not text:
        return "", set()
    
    s = str(text)
    # 1. 提取所有中括号内的内容
    brackets = re.findall(r'\[(.*?)\]', s)
    # 过滤掉特殊的后缀标签
    filtered_brackets = {'['+b+']' for b in brackets if b not in ["C", "FC", "C/FC"]}
    sorted_brackets = sorted(filtered_brackets)

    # 2. 提取小括号中的内容
    parentheses = re.findall(r'\((.*?)\)', text)
    #2.5 去除小括号中的内容
    s_cleaned = re.sub(r'\(.*?\)', '', text)

    # 3. 去除中括号及内容
    s_cleaned = re.sub(r'\[.*?\]', '', s_cleaned)
    # 4. 去除空格
    s_cleaned = s_cleaned.replace(" ", "")

    found_suffix = None
    
    # 1. 检查后缀
    if s.endswith("[C/FC]"):
        found_suffix = "[C/FC]"
    elif s.endswith("[FC]"):
        found_suffix = "[FC]"
    elif s.endswith("[C]"):
        found_suffix = "[C]"
    else:
        found_suffix = "" # 无后缀

    return s_cleaned+''.join(sorted_brackets),found_suffix,parentheses


def safe_int(val):
    """安全转换为整数，处理 None, '', 浮点数字符串等"""
    if val is None: return 0
    try:
        if isinstance(val, str):
            # 处理可能的 "10.0" 这种字符串
            return int(float(val))
        return int(val)
    except:
        return 0


def C_FC_to_E_F(item):
    if item == 'C' : return 'E'
    if item == 'FC' : return 'F'
    else: return None


def tencent_sort_key(text):
    if not text:
        return ""
    return str(text).lower()


# ================= Excel 处理类 =================

def unblock_file(file_path):
    """利用删除 NTFS 附加数据流的方式解除 Windows 文件锁定"""
    # 构造隐藏数据流的路径
    ads_path = file_path + ":Zone.Identifier"
    try:
        # 尝试删除该数据流，删除后文件就解锁了
        os.remove(ads_path)
        print(f"成功解除网络文件锁定")
    except FileNotFoundError:
        # 如果找不到该数据流，说明文件本来就没有被锁定
        pass 
    except Exception as e:
        print(f"解除锁定失败: {e}")

class ExcelProcessor:
    def __init__(self, file_path):
        self.file_path = str(Path(file_path).resolve())
        self.app = None
        self.wb = None

    def _clear_cache(self):
        try:
            temp_dir = Path(os.environ.get('TEMP', 'C:\\Temp')) / 'gen_py'
            if temp_dir.exists():
                print(f"清理 win32com 缓存: {temp_dir}")
                shutil.rmtree(temp_dir, ignore_errors=True)
        except Exception as e:
            pass

    def __enter__(self):
        self._clear_cache()
        pythoncom.CoInitialize()
        
        # 优化1: 使用 DispatchEx 创建完全独立的干净进程
        self.app = win32.DispatchEx('Excel.Application') 
        self.app.Visible = False
        # 优化2: 禁用所有弹窗警告（如：受保护的视图、更新链接等）
        self.app.DisplayAlerts = False 
        self.app.ScreenUpdating = False
        # 禁用交互，防止跳出
        self.app.Interactive = False 
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.wb:
            try: self.wb.Close(SaveChanges=False)
            except: pass
        if self.app:
            try: 
                self.app.Interactive = True # 退出前恢复交互
                self.app.DisplayAlerts = True
                self.app.Quit()
            except: pass
    


    def read_data(self):
        def read_data_main():
            # ================================================================
            # 优化3: 以只读模式打开，开启 CorruptLoad=1 让 Excel 遇到损坏直接强行修复
            self.wb = self.app.Workbooks.Open(self.file_path, UpdateLinks=0, ReadOnly=True, CorruptLoad=1)
            results = {}

            for sheet in self.wb.Sheets:
                sheet_name = sheet.Name
                # 优化4: 将 try-except 移到循环内部。这样哪怕一张表报错，也会继续读下一张表
                try: 
                    used_range = sheet.UsedRange
                    if not used_range:
                        continue

                    nrows = used_range.Rows.Count
                    ncols = used_range.Columns.Count
                    
                    raw_values = used_range.Value
                    raw_formulas = used_range.Formula
                    
                    # 优化5: 彻底解决 win32com 烦人的维度问题
                    if raw_values is None:
                        raw_values = [[None] * ncols for _ in range(nrows)]
                    elif not isinstance(raw_values, (list, tuple)):
                        # 只有1个单元格
                        raw_values = ((raw_values,),)
                    elif not isinstance(raw_values[0], (list, tuple)):
                        # 是一维元组 (多发生于只有1行或1列的情况)
                        if nrows == 1:
                            raw_values = (raw_values,)
                        else:
                            raw_values = tuple((v,) for v in raw_values)

                    if raw_formulas is None:
                        raw_formulas = [[None] * ncols for _ in range(nrows)]
                    elif not isinstance(raw_formulas, (list, tuple)):
                        raw_formulas = ((raw_formulas,),)
                    elif not isinstance(raw_formulas[0], (list, tuple)):
                        if nrows == 1:
                            raw_formulas = (raw_formulas,)
                        else:
                            raw_formulas = tuple((v,) for v in raw_formulas)

                    # 提取超链接和备注
                    hyperlinks_dict = {}
                    start_row = used_range.Row
                    start_col = used_range.Column
                    try:
                        for hl in sheet.Hyperlinks:
                            r_idx = hl.Range.Row - start_row + 1
                            c_idx = hl.Range.Column - start_col + 1
                            hyperlinks_dict[(r_idx, c_idx)] = hl.Address
                    except Exception:
                        pass

                    comments_dict = {}
                    try:
                        for cmt in sheet.Comments:
                            r_idx = cmt.Parent.Row - start_row + 1
                            c_idx = cmt.Parent.Column - start_col + 1
                            comments_dict[(r_idx, c_idx)] = cmt.Text()
                    except Exception:
                        pass

                    # 内存处理阶段
                    sheet_data = [None] * start_row
                    
                    for i in range(nrows):
                        row_data = [{}] * start_col
                        for j in range(ncols):
                            val = raw_values[i][j]
                            formula = raw_formulas[i][j]
                            
                            if val is not None:
                                if type(val).__name__ == 'time' or isinstance(val, datetime):
                                    try: val = val.isoformat()
                                    except: val = str(val)
                                elif isinstance(val, float):
                                    if val.is_integer(): val = str(int(val))
                                    else: val = str(val)

                            formula_val = formula if isinstance(formula, str) and formula.startswith('=') else None
                            hyperlink = hyperlinks_dict.get((i+1, j+1))
                            comment = comments_dict.get((i+1, j+1))

                            row_data.append({
                                'row': start_row+i,
                                'column': start_col+j,
                                'value': val,
                                'formula': formula_val,
                                'hyperlink': hyperlink,
                                'comment': comment
                            })
                            
                        sheet_data.append(row_data)
                    results[sheet_name] = sheet_data
                    print(f"成功读取工作表: {sheet_name}")

                except Exception as e:
                    # 如果某一张表读取失败，打印错误，但不会中断其他表的读取
                    print(f"⚠️ 读取工作表 '{sheet_name}' 出错，已跳过。错误详情: {e}")
                    continue

            print() 
            return results


        try:
            return read_data_main()
        except Exception as e:
            print(f"读取表格遇到错误，正在尝试修复: {e}")


            unblock_file(self.file_path)

            # ==================== 终极文件净化流水线 ====================
            temp_path = self.file_path + ".temp.zip"
            try:
                with zipfile.ZipFile(self.file_path, 'r') as zin:
                    with zipfile.ZipFile(temp_path, 'w') as zout:
                        for item in zin.infolist():
                            content = zin.read(item.filename)
                            if item.filename == 'xl/styles.xml':
                                text = content.decode('utf-8')
                                # 暴力替换所有非法的空 fill 标签，化解 openpyxl 崩溃
                                text = re.sub(r'<fill\s*>\s*</fill>', '<fill><patternFill patternType="none"/></fill>', text)
                                text = re.sub(r'<fill\s*/>', '<fill><patternFill patternType="none"/></fill>', text)
                                zout.writestr(item, text.encode('utf-8'))
                            else:
                                zout.writestr(item, content)
                os.remove(self.file_path)
                os.rename(temp_path, self.file_path)
                print("修复表格样式成功")
            except Exception as e:
                print(f"文档非法样式修复遇到问题: {e}")
                if os.path.exists(temp_path): os.remove(temp_path)

            try:
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    # 此时 openpyxl 不会再崩溃，它会完美重构文件，保全所有批注！
                    wb_fix = openpyxl.load_workbook(self.file_path)
                    wb_fix.save(self.file_path)
                    wb_fix.close()
                print("修复表格结构成功\n")
            except Exception as e:
                print(f"openpyxl修复文件失败，将尝试强行打开: {e}")
            # ============================================================

            return read_data_main()


        

# ================= 数据存储 =================

def manage_pickle(data=None, mode='save'):
    path = path = Path(BASE_DIR) / PICKLE_FILENAME
    try:
        if mode == 'save':
            with open(path, 'wb') as f:
                pickle.dump(data, f)
            return True
        else:
            if not path.exists():
                print(f"文件不存在: {path}")
                return None
            with open(path, 'rb') as f:
                return pickle.load(f)
    except Exception as e:
        print(f"Pickle操作出错 ({mode}): {e}")
        return None

# ================= 核心检查逻辑 =================


def check_cell_values_players(results):
    # 1. 建立参考库
    ref_values = {}
    ref_sheet = results.get("对照用表格", [])
    ref_no_links = set()
    
    for i in range(2, len(ref_sheet)):
        row = ref_sheet[i]
        if len(row) > 3 and row[2] and row[2].get('value'):
            clean_name = clean_player_name(row[2]['value'])
            if clean_name:
                has_link = bool(row[3].get('hyperlink'))
                ref_values[clean_name] = {
                    'has_hyperlink': has_link,
                    'hyperlink': row[3].get('hyperlink'),
                    'bilibili_name': row[3].get('value')
                }
                if not has_link:
                    ref_no_links.add(clean_name)



    all_checks = []
    unmatched = []
    cells_no_link = []
    duplicates = [] 
    total_challenges = 0

    for cfg in CHECK_CONFIGS:
        sheet_name = cfg['sheet']
        if sheet_name not in results: continue
        
        data = results[sheet_name]
        start_r, start_c = cfg['r'], cfg['c']
        
        row_map = {} 
        col_map = {}

        for i in range(start_r, len(data)):
            row_data = data[i]
            if not row_data: continue
            
            cols_to_check = [start_c] if cfg['mode'] == 'col_only' else range(start_c, len(row_data))
            
            for j in cols_to_check:
                if j >= len(row_data): continue
                cell = row_data[j]
                raw_val = cell.get('value')
                
                if not raw_val: continue
                
                val = clean_player_name(raw_val)
                if not val: continue
                
                total_challenges += 1
                has_link = bool(cell.get('hyperlink'))
                
                item = {
                    'sheet': sheet_name, 'row': i, 'col': j,
                    'cleaned': val, 'has_hyperlink': has_link,
                    'hyperlink': cell.get('hyperlink'),
                    'reference_has_hyperlink': False
                }
                
                if cfg['mode'] == 'matrix':
                    if i not in row_map: row_map[i] = {}
                    if val not in row_map[i]: row_map[i][val] = []
                    row_map[i][val].append(j)
                else:
                    if val not in col_map: col_map[val] = []
                    col_map[val].append(i)

                if val in ref_values:
                    ref_info = ref_values[val]
                    item['reference_has_hyperlink'] = ref_info['has_hyperlink']
                    item['reference_hyperlink'] = ref_info['hyperlink']
                    if not has_link:
                        cells_no_link.append(item)
                    all_checks.append(item) 
                else:
                    unmatched.append(item)

        if cfg['mode'] == 'matrix':
            for r, p_map in row_map.items():
                for p, cols in p_map.items():
                    if len(cols) > 1:
                        cols_str = " ".join([f"列{int_to_excel_column(c)}" for c in cols])
                        duplicates.append(f"  '{sheet_name}'第{r}行: 玩家 '{p}' 在 {cols_str}")
        else:
            for p, rows in col_map.items():
                if len(rows) > 1:
                    rows_str = " ".join([f"行{r}" for r in rows])
                    duplicates.append(f"  '{sheet_name}': 玩家 '{p}' 在 {rows_str}")

    print(f"玩家总数: {len(ref_values)+1}, 挑战总数: {total_challenges}\n")
    
    if ref_no_links:
        print("对照表中无超链接的玩家:")
        for idx, p in enumerate(sorted(ref_no_links), 1):
            print(f"  {idx:2d}. {p}")
    else: print("对照表玩家均附有超链接！")
    print()

    if cells_no_link:
        print("挑战单元格中无超链接的项目:")
        for idx, c in enumerate(cells_no_link, 1):
            print(f"  {idx:2d}. {c['sheet']} - 行{c['row']}列{int_to_excel_column(c['col'])}: {c['cleaned']}")
    else: print("挑战单元格中均附有超链接！")
    print()

    if duplicates:
        print("重复登记检测结果:")
        print(*duplicates, sep='\n')
    else: print("未发现重复登记的情况！")
    print()

    if unmatched:
        print("未在对照表中找到的单元格:")
        for idx, item in enumerate(unmatched, 1):
            print(f"  {idx:2d}. {item['sheet']} - 行{item['row']}列{int_to_excel_column(item['col'])}: {item['cleaned']}")
            pass
    else: print("所有单元格的玩家都在对照表中找到了！")
    print()

    return [all_checks, ref_values]

# ================= 核心检查逻辑 =================

def check_column_sorting(col_idx, tier_name, items, errors):
    """
    检查一列数据的排序情况
    区分：完全乱序 vs Subtier内部混入错误难度
    """
    if not items: return

    # 1. 提取 Subtier (High/Mid/Low)
    is_split_tier = (tier_name in IS_SUB_TIERED)

    
    # === 非分层 Tier (T4+) 的简单检查 ===
    if not is_split_tier:
        names = [x['name'] for x in items]
        # 忽略大小写进行比较
        sorted_names = sorted(names, key=tencent_sort_key)
        if names != sorted_names:
            errors.append(f"主榜列 {tier_name}: 列表未按字典序排序")
        return

    # === 分层 Tier (T0-T3) 的复杂检查 ===
    current_list = items
    
    # 特殊处理: High T0 (手动提取)
    if tier_name == "T0" and HIGH_T0_CHALLENGES:
        high_part = []
        rem_part = []
        for it in current_list:
            # 简单匹配: 只要名字出现在白名单里
            if it['name'] in HIGH_T0_CHALLENGES:
                high_part.append(it)
            else:
                rem_part.append(it)
        
        # 检查 High T0 内部排序
        h_names = [x['name'] for x in high_part]
        if h_names != sorted(h_names, key=tencent_sort_key):
             errors.append(f"主榜列 {tier_name} (High T0): 未按字典序排序")
        
        # 继续检查剩余部分 (Mid/Low)
        current_list = rem_part
    
    if not current_list: return

    # 2. 自动分层 (利用字母序断层)
    # 算法: 只要发现 next < prev，就视为新的一层
    layers = []
    current_layer = [current_list[0]]
    
    for i in range(1, len(current_list)):
        curr = current_list[i]
        prev = current_list[i-1]
        # 忽略大小写比较
        if tencent_sort_key(curr['name']) < tencent_sort_key(prev['name']):
            layers.append(current_layer)
            current_layer = []
        current_layer.append(curr)
    layers.append(current_layer)
    
    # === 情况 1: 无法分成合理的 Subtier (层数过多) ===
    # T0 如果提取了 High，剩下应该只有 Mid/Low (2层)
    # T1-T3 应该是 High/Mid/Low (3层)
    # 放宽一点限制，只要超过3层，肯定就是乱序了
    if len(layers) > 3:
        errors.append(f"主榜列 {tier_name}: 似乎有排序错误")
        return

    # === 情况 2: 可以分层，但在层内混入了其他难度的挑战 ===
    tier_order_map = {"High": 0, "Mid": 1, "Low": 2, "Unknown": 3}
    last_order = -1
    
    for i, layer in enumerate(layers):
        if not layer: continue
        
        # 统计该层内最多的难度标签 (确定这一层本来应该是哪个 Subtier)
        tag_counts = {"High": 0, "Mid": 0, "Low": 0}
        for it in layer:
            d = it['diff']
            if "High" in d: tag_counts["High"] += 1
            elif "Mid" in d: tag_counts["Mid"] += 1
            elif "Low" in d: tag_counts["Low"] += 1
        
        # 找出主导标签
        main_tag = max(tag_counts, key=tag_counts.get)
        current_order = tier_order_map.get(main_tag, 3)
        
        # 检查层内混杂
        for it in layer:
            # 如果该挑战的难度不包含主导标签
            if main_tag not in it['diff']:
                errors.append(f"主榜列 {tier_name}: 挑战 '{it['name']}' (实际难度 {it['diff']}) 混入了 {main_tag} 区域")

        # 检查层级顺序 (例如不能先 Low 再 High)
        if current_order < last_order:
             errors.append(f"主榜列 {tier_name}: Subtier 顺序错误，发现 {main_tag} 出现在更高级别之后")
        last_order = current_order


def get_map_configs(results,source_sheets):
    """
    读取详细信息，构建地图配置库。
    包含难度配置和通过人数。
    """
    map_db = {}
    map_db_row = {}
    
    for sheet_name in source_sheets:
        if sheet_name not in results: continue
        
        data = results[sheet_name]
        # 从第3行开始
        for i in range(3, len(data)):
            row = data[i]
            if len(row) < 3: continue

            pack_name = row[1].get('value') if len(row) > 1 else ""
            raw_map_name = row[2].get('value') if len(row) > 2 else ""
            if not raw_map_name: raw_map_name = pack_name
            if not raw_map_name: continue

            j = 0
            pack_name = ''
            if not pack_name:
                for j in range(i):
                    row_j = data[i-j]
                    pack_name = row_j[1].get('value') if len(row_j) > 1 else ""
                    if pack_name:
                        break
            total_name = pack_name + ' - ' + raw_map_name
            
            clean_name,t1,t2 = clean_map_name_this_is_a_function(raw_map_name)
            if not clean_name: continue
            
            diff_c = row[3].get('value')
            diff_fc = row[4].get('value')
            
            # 读取人数
            count_total = safe_int(row[5].get('value')) if len(row) > 5 else 0
            count_fc = safe_int(row[6].get('value')) if len(row) > 6 else 0
            formula_total = row[5].get('formula')
            formula_fc = row[6].get('formula')



            term = {
                'diff_c': diff_c,
                'diff_fc': diff_fc,
                'count_total': count_total,
                'count_fc': count_fc,
                'formula_total': formula_total,
                'formula_fc': formula_fc,
                'row': i,
                'sheet_name': sheet_name,
                'clean_name': clean_name,
                'source': f"{sheet_name} 行{i}",
                'total_name': total_name
            }

            if map_db.get(clean_name):
                map_db[clean_name].append(term)
            else:
                map_db[clean_name] = [term]
                

            map_db_row[i] = term
        
    return map_db,map_db_row





def challenge_count_is_right_ma(map_db_row):

    errors =[]
    for _ , map_info in map_db_row.items():
        diff_c = map_info['diff_c']
        diff_fc = map_info['diff_fc']

        f_total = map_info['formula_total']
        f_fc = map_info['formula_fc']
        f = {'C':f_total,
             'FC':f_fc  }

        row = map_info['row']
        sheet_name = map_info['sheet_name']

        # 规则 1: FC格为 "-"
        if diff_fc == "-":
            if f_fc:
                errors.append(f"'{sheet_name}'第{row}行'FC'列的达成人数公式应该为空，而不是{f_fc}")
            a = challenge_count_is_right_ma_fuzhuhanshu(f, 'C', row, sheet_name)
            if a:
                errors.append(a)
            

        # 规则 2: FC格为空
        elif not diff_fc:
            a = challenge_count_is_right_ma_fuzhuhanshu(f, 'C and FC', row, sheet_name)
            if a:
                errors.append(a)

        # 规则 3: FC格存在难度
        else:
            if diff_fc in DIFFICULTY_TYPES and diff_c == "-":
                if f_total:
                    errors.append(f"'{sheet_name}'第{row}行'C'列的达成人数公式应该为空，而不是{f_total}")
                a = challenge_count_is_right_ma_fuzhuhanshu(f, 'FC', row, sheet_name)
                if a:
                    errors.append(a)

            if diff_c and diff_c in DIFFICULTY_TYPES and diff_c != "-":
                a = challenge_count_is_right_ma_fuzhuhanshu(f, 'C and FC', row, sheet_name)
                if a:
                    errors.append(a)
        
    return errors



def challenge_count_is_right_ma_fuzhuhanshu(f, is_c_fc, row, sheet_name):

    row_str = str(row)
    pattern = None
    f_ = ''
    
    if is_c_fc == 'C':
        # (?:[A-Z]{2,}) 确保列号是三个及以上字母
        pattern = rf'^=COUNTA\(G{row_str}:(?:[A-Z]{{2,}}){row_str}\)$'
        f_= f.get('C','')
        
    elif is_c_fc == 'FC':
        # 匹配模式：=COUNTIF(G<行号>:<较大列号><行号>,"*[FC]*")
        pattern = rf'^=COUNTIF\(G{row_str}:(?:[A-Z]{{2,}}){row_str},["]\*\[FC\]\*["]\)$'
        f_= f.get('FC','')

    if pattern:
        if isinstance(f_, str) and bool(re.match(pattern, f_)):
            return ''
        else:
            return f"'{sheet_name}'第{row}行'{is_c_fc}'列的达成人数公式'{f_}'有误"

    if is_c_fc == 'C and FC':
        c_msg = challenge_count_is_right_ma_fuzhuhanshu(f,'C', row, sheet_name)
        fc_msg = challenge_count_is_right_ma_fuzhuhanshu(f,'FC', row, sheet_name)
        messages = []
        if c_msg: 
            messages.append(c_msg)
        if fc_msg: 
            messages.append(fc_msg)

        return "\n".join(messages)

    return ''
        


def validate_challenge_and_count(suffix, map_info):
    """
    校验后缀合法性，并返回其实际难度和人数。
    返回: (is_valid, real_diff, real_count, msg , is_fc_equals_c, challenge_count_errors)
    """
    diff_c = map_info['diff_c']
    diff_fc = map_info['diff_fc']

    c_total = map_info['count_total']
    c_fc = map_info['count_fc']

    
    # 规则 1: FC格为 "-"
    if diff_fc == "-":
        if suffix == "":
            # 人数就是总Clear数
            return True, diff_c, c_total, "OK",False
        else:
            return False, None, 0, f"不应有后缀",False

    # 规则 2: FC格为空
    elif not diff_fc:
        if suffix == "[C/FC]":
            # 人数是总Clear数
            return True, diff_c, c_total, "OK",False
        else:
            return False, None, 0, f"后缀应为 [C/FC]",False

    # 规则 3: FC格存在难度
    else:
        if suffix == "[FC]":
            if diff_fc in DIFFICULTY_TYPES:
                # [FC] 挑战的人数就是 FC数
                return True, diff_fc, c_fc, "OK",False
            else:
                return False, None, 0, f" FC 难度不存在",False
        
        elif suffix == "[C]":
            if diff_c and diff_c in DIFFICULTY_TYPES and diff_c != "-":
                # [C] 挑战的人数 = 总Clear数 - FC数 (仅Clear)
                c_only = c_total - c_fc
                if c_only == 0:
                    return True, diff_c, c_only, "OK",True
                return True, diff_c, c_only, "OK",False
            else:
                return False, None, 0, f" Clear 难度不存在",False
        
        elif suffix == "[C/FC]":
             return False, None, 0, "不应使用 [C/FC]",False
        else:
             return False, None, 0, f"应为 [FC]或[C]",False



def is_every_element_a_substring(total_name,pa):
    if total_name:
        ret = True
        for papapa in pa:
            if papapa not in total_name:
                ret = False
        return ret
    else:
        return False


def find_map_config_by_v_and_pa(might_be_map_config_set,pa):

    #比对逻辑：pa的每一项必须出现在pack_name+map_name这个串里
    map_config = None
    if might_be_map_config_set:
        for might_be_map_config in might_be_map_config_set:
            might_be_total_name = might_be_map_config.get('total_name')
            if is_every_element_a_substring(might_be_total_name,pa):
                map_config = might_be_map_config
                
    return map_config




def main_zhubang_check(results):
    errors = []

    correct_zhubang = []
    
    # 1. 预先构建详细信息库
    map_db , map_db_row = get_map_configs(results,["详细信息"])

    errors.extend(challenge_count_is_right_ma(map_db_row))

    _ , map_db_row_other_challenges = get_map_configs(results,["其他挑战"])
    errors.extend(challenge_count_is_right_ma(map_db_row_other_challenges))


    if not map_db: return ["错误：无法构建地图数据库"]

    main_sheet = results.get("主榜")
    detail_sheet = results.get("详细信息")
    
    if not main_sheet or not detail_sheet: 
        return ["缺少主榜或详细信息数据"]
    
    found_challenges = []
    tier_counts = {k: 0 for k in DIFFICULTY_TYPES}
    
    # 定义检查列和对应的 Tier
    zhubang_name_to_row = {}
    # 遍历需要检查的列
    for col_idx in CHECK_COLUMNS_MAIN:
        if len(main_sheet) > 1 and col_idx >= len(main_sheet[1]): 
            break

        current_tier_group = TIER_MAP_MAIN_REV.get(col_idx, "Unknown")
        allowed_diffs = []
        for t_name, t_list in TIER_GROUPS:
            if t_name == current_tier_group:
                allowed_diffs = t_list
                break

        col_items = []

        # 遍历行 (从第4行开始)
        for r in range(4, len(main_sheet)):
            row = main_sheet[r]
            if col_idx >= len(row): continue

            # === 获取主榜单元格数据 ===
            cell = row[col_idx]
            val = cell.get('value')

            # 跳过空值
            if not val or not str(val):
                continue

            loc_str = f"主榜 {current_tier_group} 列第{r}行"

            # =========================================================
            # PART 1: 严格集成 check_formula_values_main 的逻辑
            # =========================================================
            
            # 1. 检查右侧单元格是否存在
            if col_idx + 1 >= len(row):
                errors.append(f"{loc_str}: 没有右边的单元格 (缺少公式)")

            right_cell = row[col_idx + 1]
            formula = right_cell.get('formula')
            
            # 2. 提取行号
            num = extract_number_from_formula(formula)
            if num is None:
                errors.append(f"{loc_str}: 公式无效或缺少引用数字")
            # 3. 检查行号范围
            if not (1 <= num <= 10000):
                errors.append(f"{loc_str}: 引用行号{num}超出范围(1-10000)")
            if num >= len(detail_sheet):
                errors.append(f"{loc_str}: 引用行号{num}超出详细信息表范围")

            # 5. 执行 clean_text_common 并比对
            v,suffix,pa = clean_map_name_this_is_a_function(val)

            #比对逻辑：pa的每一项必须出现在pack_name+map_name这个串里
            map_config = find_map_config_by_v_and_pa(map_db.get(v),pa)
            

            name_which_the_num_corresponds = map_db_row.get(num,{}).get('clean_name')

            if not map_config:
                errors.append(f"{loc_str} 未在详细信息找到")
                continue

            num_should_be = map_config.get('row')
            if num_should_be != num:
                errors.append(f"{loc_str} {v} 不对应详细信息第{num}行 {name_which_the_num_corresponds} , 它应该对应行 {num_should_be}")

            # 校验挑战逻辑 (后缀是否正确，难度是否存在，获取实际难度和人数)
            is_valid, real_diff, real_count, msg ,is_fc_equals_c = validate_challenge_and_count(suffix, map_config)
            
            if not is_valid:
                errors.append(f"{loc_str} ('{val}'): {msg}")

            if is_valid:
                # 检查人数 > 0
                if not is_fc_equals_c:
                    if real_count <= 0:
                        errors.append(f"{loc_str} ('{val}'): 达成人数为 {real_count}，不应出现在榜单")

                # 检查难度是否匹配当前列
                if real_diff not in allowed_diffs:
                    errors.append(f"{loc_str} ('{val}'): 实际难度 '{real_diff}' 不匹配此列 ({current_tier_group})")
                
                # 记录数据用于覆盖率和排序
                c_name = map_config.get('clean_name')
                id = str(val)
                if id in found_challenges:
                    errors.append(f"挑战 {id} 在主榜出现多次")
                else:   
                    found_challenges.append(id)
                    tier_counts[real_diff] += 1

                col_items.append({
                    'name': id, # 使用主榜原始显示文本
                    'diff': real_diff
                })



        # 列内排序检查
        if col_items:
            check_column_sorting(col_idx, current_tier_group, col_items, errors)

    # =========================================================
    # PART 3: 覆盖率检查 (检查遗漏)
    # =========================================================
    errors_coverge , correct_zhubang , can_use_for_zhubang = check_coverage(map_db_row, found_challenges)
    errors.extend(errors_coverge)

    if errors:
        print(*errors, sep='\n')
        print()
        if can_use_for_zhubang:
            try:
                generate_correct_zhubang(correct_zhubang)
            except Exception as e:
                print(f'\n导出正确的主榜时出错:{e}')
        else:
            print('\n目前主榜登记的挑战名称有遗漏，无法获取名称以导出正确的主榜')
    else:
        print("挑战难度和达成人数检查通过！")
    return None


def check_coverage(map_db_row, found_ids):
    """
    检查是否所有 有效人数的 挑战都在主榜出现了
    通过 clean_name 加上 pa 共同验证
    """
    errors_coverge = []
    correct_zhubang = []
    can_use_for_zhubang = True

    for row, info in map_db_row.items():
        diff_c = info['diff_c']
        diff_fc = info['diff_fc']
        c_total = info['count_total']
        c_fc = info['count_fc']
        row = info['row']
        clean_name = info['clean_name']
        total_name = info['total_name'] # 提取完整的名称以供 pa 匹配

        expected_items = []

        # 生成该行地图应该存在的各个挑战形态：(期望的clean_name, 期望的后缀, 难度, 人数, 行号, C/FC标志)
        if diff_fc == "-":
            if diff_c in DIFFICULTY_TYPES and c_total > 0:
                expected_items.append((total_name, clean_name, "", diff_c, c_total, row, 'C'))
        
        elif not diff_fc:
            if diff_c in DIFFICULTY_TYPES and c_total > 0:
                expected_items.append((total_name, clean_name, "[C/FC]", diff_c, c_total, row, 'C'))

        else:
            if diff_fc in DIFFICULTY_TYPES and c_fc > 0:
                expected_items.append((total_name, clean_name, "[FC]", diff_fc, c_fc, row, 'FC'))
            if diff_c in DIFFICULTY_TYPES and diff_c != "-" and (c_total - c_fc) > 0:
                expected_items.append((total_name, clean_name, "[C]", diff_c, c_total - c_fc, row, 'C'))
                
        # 遍历核对每个期望的挑战是否在主榜出现
        for e_total_name, e_clean_name, e_suffix, diff, count, row, col in expected_items:
            
            match_found = False
            real_name = None

            # 遍历主榜找到的所有挑战
            for f_id in found_ids:

                # 重新解析主榜上的名称以获取它的 clean_name , suffix 和 pa
                f_clean_name, f_suffix, f_pa = clean_map_name_this_is_a_function(f_id)

                # 验证规则 1：clean_name 必须相等，且挑战后缀([FC]等)必须一致
                if f_clean_name == e_clean_name:
                    if f_suffix == e_suffix:
  
                        # 验证规则 2：pa 里的所有元素必须都是 total_name 的子串
                        if is_every_element_a_substring(total_name,f_pa):
                            match_found = True
                            break
            
            real_name = f_id

            # 构建用于报错信息显示的组合名称
            eid = f"{e_total_name}{e_suffix}"
            
            if not match_found:
                errors_coverge.append(f"遗漏: 挑战 '{eid}' (难度 {diff}, 人数 {count}) 未在主榜出现")
                can_use_for_zhubang = False
            else:
                correct_zhubang.append({
                    'name': real_name,
                    'diff': diff,
                    'row' : row,
                    'col' : col  
                })

    return errors_coverge, correct_zhubang, can_use_for_zhubang







def generate_correct_zhubang(correct_zhubang):
    """
    将主榜数据导出到 Excel 文件。
    """
    main_bang_path = os.path.join(BASE_DIR, MAIN_BANG_FILENAME)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "主榜"

    # --- 新增：定义统一的字体和对齐样式 ---
    cambria_14_font = Font(name='Cambria', size=14)
    left_alignment = Alignment(horizontal='left')
    right_alignment = Alignment(horizontal='right')

    # 1. 按照“列号”将数据分组
    # column_data 结构: { col_num: [item1, item2, ...] }
    column_data = defaultdict(list)
    
    for item in correct_zhubang:
        diff = item['diff']
        if diff in TIER_MAP_MAIN:
            col_num, color = TIER_MAP_MAIN[diff]
            column_data[col_num].append(item)
            
    # 2. 遍历每个奇数列组，执行排序并写入数据
    for col_num, items in column_data.items():
        # 使用 9999 作为 fallback 容错处理未知难度
        sorted_items = sorted(
            items,
            key=lambda x: (
                DIFFICULTY_TYPES.index(x['diff']) if x['diff'] in DIFFICULTY_TYPES else 9999,
                tencent_sort_key(x['name'])
            )
        )

        # 3. 逐行写入排好序的数据
        for row_idx, item in enumerate(sorted_items, start=1):
            diff = item['diff']
            _, color_code = TIER_MAP_MAIN[diff]
            
            # 确保颜色代码没有 '#' 前缀，openpyxl 要求纯 ARGB/RGB HEX 格式 (如 'FF0000')
            clean_color = str(color_code).replace('#', '')
            fill_style = PatternFill(start_color=clean_color, end_color=clean_color, fill_type="solid")
            
            # --- 奇数列 (地图名称) ---
            cell_left = ws.cell(row=row_idx, column=col_num)
            cell_left.value = item['name']
            cell_left.fill = fill_style
            cell_left.font = cambria_14_font       # 设置为 Cambria, 14号
            cell_left.alignment = left_alignment   # 设置为左对齐

            # --- 偶数列 (达成人数，紧跟在其右边) ---
            cell_right = ws.cell(row=row_idx, column=col_num + 1)
            # Excel 绝对引用格式：='表名'!$列$行 (假设字典中 col='A', row=2)
            cell_right.value = f"='详细信息'!${C_FC_to_E_F(item['col'])}${item['row']}"
            cell_right.fill = fill_style
            cell_right.font = cambria_14_font      # 设置为 Cambria, 14号
            cell_right.alignment = right_alignment # 设置为右对齐
            
    # 4. 保存文件
    wb.save(main_bang_path)
    print(f"自动生成的主榜文件已成功导出到: {main_bang_path}")

    return None






def export_hyperlinks(data_pack, results):
    cells, players = data_pack
    path = Path(BASE_DIR) / HYPERLINK_FILENAME

    try:
        # 直接读取"对照用表格"全表中的超链接
        ref_sheet = results.get("对照用表格", [])
        p_links = []
        for i in range(len(ref_sheet)):
            row_data = ref_sheet[i]
            if not row_data:
                continue
            for j in range(len(row_data)):
                cell = row_data[j]
                if cell and cell.get('hyperlink') and cell.get('value'):
                    r = cell.get('row', i)
                    c = cell.get('column', j)
                    p_links.append(
                        f"{cell['hyperlink']} [对照用表格 行{r} 列{int_to_excel_column(c)} 原B站名'{cell['value']}']"
                    )
        
        c_links = [
            f"{c['hyperlink']} [{c['sheet']} 行{c['row']} 列{int_to_excel_column(c['col'])} 玩家'{c.get('cleaned','')}']"
            for c in cells
        ]

        with open(path, 'w', encoding='utf-8') as f:
            if p_links:
                f.write("=== 玩家链接 ===\n")
                f.write("\n".join(p_links) + "\n\n")
            if c_links:
                f.write("=== 单元格链接 ===\n")
                f.write("\n".join(c_links) + "\n")

        print(f"超链接数据已写入: {path}\n")
    except Exception as e:
        print(f"写入文件出错: {e}")






def rename_backup(file_path):
    path = Path(file_path)
    if not path.exists():
        print(f"错误：文件 '{path}' 不存在！")
        return
    ts = datetime.now().strftime("%Y年%m月%d日 %H.%M")
    new_name = f"{path.stem} 备份 {ts}{path.suffix}"
    new_path = path.parent / new_name
    try:
        path.rename(new_path)
        return new_path,new_name
    except Exception as e:
        print(f"备份出错: {e}")






def backup_file(copy_file_path,new_path,new_name):
    if not copy_file_path: return
    copy_path = Path(copy_file_path)
    if not copy_path.exists():
        print(f"错误：路径 '{copy_path}' 不存在！")
        return
    destination_path = copy_path / new_name
    try:
        shutil.copy2(new_path, destination_path)
        print(f"文件已成功备份到: {destination_path}")
    except Exception as e:
        print(f"备份出错: {e}")




# ================= 主程序 =================


def main():

    # 1. 读取表格数据
    if DO_READ_FILE:
        with ExcelProcessor(FILE_PATH) as proc:
            data = proc.read_data()
            if data: manage_pickle(data, 'save')

    # 2. 加载数据
    results = manage_pickle(mode='load')

    if results:
        # 3. 执行检查
        hyperlink_data = check_cell_values_players(results)
        main_zhubang_check(results)

        export_hyperlinks(hyperlink_data, results)

        # 4. 备份文件
        rename_backup_FILE_PATH = rename_backup(FILE_PATH)
        if rename_backup_FILE_PATH:
            new_path,new_name = rename_backup_FILE_PATH
    
        while(True):
            response = input("输入 Y 备份文件并结束，输入 R 重新运行，输入 E 直接结束").strip().upper()

            if response == 'Y':
                for copy_file_path in COPY_FILE_PATH_LIST:
                    backup_file(copy_file_path,new_path,new_name)
                return

            if response == 'R':
                main()
                return

            if response == 'E':
                return

            else: continue

if __name__ == "__main__":
    main()
    input("\n程序运行结束，按回车键退出...")

    
        